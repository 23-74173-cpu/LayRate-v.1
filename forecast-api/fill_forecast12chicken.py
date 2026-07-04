"""Fill forecast12chicken.xlsx using daily data from EggRecord_2026_updated.xlsx.

This script reads the raw daily production sheets from EggRecord, aggregates
per-cage totals, and writes the matching daily values into the editable columns
of forecast12chicken.xlsx (Egg_Count, Temperature_C, Humidity_Percent,
Mortality_Count, plus sample feed values).
"""

import re
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl.styles import Protection


EGG_RECORD_PATH = Path(__file__).resolve().parent / "EggRecord_2026_updated.xlsx"
FORECAST_PATH = Path(__file__).resolve().parent / "forecast12chicken.xlsx"


def parse_date_from_label(label: str, default_year: int = 2026) -> date | None:
    """Parse strings like 'Apr 24 (Thu)' into a date."""
    if pd.isna(label):
        return None
    label = str(label).strip()
    if not label:
        return None
    # Extract the month/day portion before the day-of-week in parentheses.
    match = re.match(r"([A-Za-z]+)\s+(\d+)", label)
    if not match:
        return None
    month_str, day_str = match.groups()
    parsed = pd.to_datetime(f"{month_str} {day_str} {default_year}", errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.date()


def extract_egg_record_data(path: Path) -> pd.DataFrame:
    """Read all production sheets and return a tidy daily DataFrame."""
    xl = pd.ExcelFile(path, engine="openpyxl")
    rows = []

    for sheet_name in xl.sheet_names:
        if sheet_name == "SUMMARY":
            continue

        df = pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl", header=None)

        # Find the header row containing 'DATE' in column 0.
        header_row = None
        for idx, row in df.iterrows():
            if str(row.get(0, "")).strip().upper().startswith("DATE"):
                header_row = idx
                break
        if header_row is None:
            continue

        data = df.iloc[header_row + 2 :].copy()  # skip sub-header row
        data.columns = df.iloc[header_row]

        for _, row in data.iterrows():
            record_date = parse_date_from_label(row.get("DATE"))
            if record_date is None:
                continue

            # CAGE 1 is columns 1-3 (R1, R2, R3), CAGE 2 is 4-6, CAGE 3 is 7-9.
            cage1_eggs = sum(pd.to_numeric(row.iloc[i], errors="coerce") for i in range(1, 4))
            cage2_eggs = sum(pd.to_numeric(row.iloc[i], errors="coerce") for i in range(4, 7))
            cage3_eggs = sum(pd.to_numeric(row.iloc[i], errors="coerce") for i in range(7, 10))

            # Fixed column positions in the EggRecord daily sheets:
            # 0=DATE, 1-3=CAGE1, 4-6=CAGE2, 7-9=CAGE3, 10=TOTAL EGGS,
            # 11=MORT., 12=STATUS, 13=TEMP Min, 14=TEMP Max, 15=HUM Min,
            # 16=HUM Max, 17=NOTES.
            temp_min = pd.to_numeric(row.iloc[13], errors="coerce")
            temp_max = pd.to_numeric(row.iloc[14], errors="coerce")
            hum_min = pd.to_numeric(row.iloc[15], errors="coerce")
            hum_max = pd.to_numeric(row.iloc[16], errors="coerce")
            mortality = pd.to_numeric(row.iloc[11], errors="coerce")

            rows.append(
                {
                    "date": record_date,
                    "cage1_eggs": int(cage1_eggs) if pd.notna(cage1_eggs) else 0,
                    "cage2_eggs": int(cage2_eggs) if pd.notna(cage2_eggs) else 0,
                    "cage3_eggs": int(cage3_eggs) if pd.notna(cage3_eggs) else 0,
                    "temperature_c": round((temp_min + temp_max) / 2, 2)
                    if pd.notna(temp_min) and pd.notna(temp_max)
                    else None,
                    "humidity_percent": round((hum_min + hum_max) / 2, 2)
                    if pd.notna(hum_min) and pd.notna(hum_max)
                    else None,
                    "mortality_count": int(mortality) if pd.notna(mortality) and mortality > 0 else 0,
                }
            )

    return pd.DataFrame(rows).sort_values("date").reset_index(drop=True)


def fill_forecast12chicken():
    source = extract_egg_record_data(EGG_RECORD_PATH)
    source = source.drop_duplicates(subset=["date"], keep="first")

    forecast = pd.read_excel(FORECAST_PATH, engine="openpyxl")
    forecast["Date"] = pd.to_datetime(forecast["Date"], errors="coerce").dt.date

    # Ensure editable columns can hold mixed types (string/numeric/None).
    for col in [
        "Egg_Count",
        "Temperature_C",
        "Humidity_Percent",
        "Feed_Batch_Code",
        "Crude_Protein_Percent",
        "Feed_Consumed_kg",
        "Mortality_Count",
    ]:
        forecast[col] = forecast[col].astype(object)

    # Map CAGE 1 data to the single CAGE-A in forecast12chicken.xlsx.
    lookup = {
        row["date"]: row
        for _, row in source.iterrows()
    }

    for idx, row in forecast.iterrows():
        d = row.get("Date")
        if d not in lookup:
            continue

        data = lookup[d]
        forecast.at[idx, "Egg_Count"] = data["cage1_eggs"]
        forecast.at[idx, "Temperature_C"] = data["temperature_c"]
        forecast.at[idx, "Humidity_Percent"] = data["humidity_percent"]
        forecast.at[idx, "Mortality_Count"] = data["mortality_count"]
        forecast.at[idx, "Feed_Batch_Code"] = "BATCH-001"
        forecast.at[idx, "Crude_Protein_Percent"] = 18.0
        # Approximate daily feed at ~0.11 kg/hen/day.
        forecast.at[idx, "Feed_Consumed_kg"] = round(row.get("Hen_Count", 48) * 0.11, 2)

    # Preserve the original sheet name and protection settings.
    from openpyxl import load_workbook

    forecast.to_excel(FORECAST_PATH, index=False, engine="openpyxl", sheet_name="Forecast Input")

    wb = load_workbook(FORECAST_PATH)
    if "Sheet1" in wb.sheetnames and "Forecast Input" not in wb.sheetnames:
        wb["Sheet1"].title = "Forecast Input"
    ws = wb["Forecast Input"]

    locked_columns = {
        "Date",
        "Cage_Code",
        "Breed",
        "Flock_Age_Weeks",
        "Hen_Count",
    }
    editable_col_indices = [
        idx for idx, col in enumerate(forecast.columns, start=1) if col not in locked_columns
    ]

    for row in range(2, ws.max_row + 1):
        for col_idx in editable_col_indices:
            ws.cell(row=row, column=col_idx).protection = Protection(locked=False)
    ws.protection.sheet = True

    wb.save(FORECAST_PATH)

    print(f"Filled {len(forecast)} row(s) in {FORECAST_PATH.name}")
    print(f"Source records matched: {len([d for d in forecast['Date'] if d in lookup])}")


if __name__ == "__main__":
    fill_forecast12chicken()
